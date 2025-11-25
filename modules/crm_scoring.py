# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import pandas as pd
from dateutil import tz
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# ====== 상수 ======
ABSOLUTE_SCORE_CUT = 65.0  # 절대 점수 컷 (0~100 스케일)

# 오늘 날짜 (Asia/Seoul)
KST = tz.gettz("Asia/Seoul")
today_dt = datetime.now(tz=KST)
today = today_dt.date()
file_prefix = today_dt.strftime("%y%m%d")  # 예: 251111


# ====== 유틸 함수 ======
def find_col(cols, keywords_list):
    """여러 후보 키워드(리스트의 리스트)를 받아, 컬럼 이름에 모두 포함되는 첫 컬럼을 찾음."""
    clower = [str(c).strip().lower() for c in cols]
    for kw_group in keywords_list:
        for i, c in enumerate(clower):
            if all(kw in c for kw in kw_group):
                return cols[i]
    return None


def to_date(s):
    if pd.isna(s):
        return pd.NaT
    try:
        return pd.to_datetime(s, errors="coerce").date()
    except Exception:
        return pd.NaT


def to_num(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        cleaned = re.sub(r"[^\d\.-]", "", x.replace(",", ""))
    else:
        cleaned = x
    try:
        return float(cleaned)
    except Exception:
        return np.nan


def extract_group_anywhere(name, first_date):
    """
    이름 어디에 있든 영문이 있으면 마지막 글자 기준으로 A/C/D/E 결정.
    없으면 최초구매 연도 규칙(2024->A, 2025->E, 그 외 A)로 결정하고 그 알파벳을 이름 끝에 붙임.
    """
    raw = "" if pd.isna(name) else str(name).strip()
    letters = re.findall(r"[A-Za-z]", raw)
    group = None
    if letters:
        last_char = letters[-1].upper()
        if last_char in {"A", "C", "D", "E"}:
            group = last_char
    if group is None:
        yr = first_date.year if (first_date is not pd.NaT and first_date is not None) else None
        if yr == 2024:
            group = "A"
        elif yr == 2025:
            group = "E"
        else:
            group = "A"
        raw = f"{raw}{group}"
    return group, raw


def over_90_days(d, ref_date):
    if pd.isna(d):
        return True
    return (ref_date - d) >= timedelta(days=90)


def weight_text(g, s100):
    # 표시용 텍스트. 점수 없으면 '-'
    if s100 is None or (isinstance(s100, float) and np.isnan(s100)):
        return f"{g}군 가중치 -"
    try:
        val = int(round(float(s100)))
    except Exception:
        val = 0
    return f"{g}군 가중치 {val}"


def compute_corr_weights(feat_df: pd.DataFrame) -> dict:
    """상관관계 기반 가중치 계산. 합이 10이 되도록 스케일."""
    if len(feat_df) < 3:
        return {"정제 총 매출": 4.0, "구매 횟수": 3.0, "평균 구매금": 3.0}  # 데이터 부족 시 폴백
    corr = feat_df.corr(method="pearson").abs()
    related = {}
    for col in corr.columns:
        others = [c for c in corr.columns if c != col]
        vals = corr.loc[others, col].dropna().values
        related[col] = float(np.mean(vals)) if len(vals) > 0 else 0.0
    s = sum(related.values())
    if s <= 1e-8:
        return {"정제 총 매출": 4.0, "구매 횟수": 3.0, "평균 구매금": 3.0}
    scaled = {k: v * 10.0 / s for k, v in related.items()}  # 합=10으로 정규화
    return scaled


def beautify_excel(path, sheet_name="대상"):
    """타깃 파일(김훈하/전정미/백인보) 서식 적용."""
    wb = load_workbook(path)
    ws = wb[sheet_name]

    # 전체 중앙 정렬
    center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center

    # 숫자 서식: 콤마/정수
    header_to_colidx = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

    def fmt(header, fmt_str):
        if header in header_to_colidx:
            c = header_to_colidx[header]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c).number_format = fmt_str

    fmt("정제 총 매출", "#,##0")
    fmt("평균 구매금", "#,##0.##")
    fmt("구매 횟수", "#,##0")
    fmt("가중치(점수)", "0")

    # 열 너비 자동 확장
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            text = "" if val is None else str(val)
            max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = max(10, min(60, int(max_len * 1.3)))

    # 자동 필터 + 헤더 고정
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # 정렬용 숫자열 숨김
    if "가중치(점수)" in header_to_colidx:
        ws.column_dimensions[get_column_letter(header_to_colidx["가중치(점수)"])].hidden = True

    wb.save(path)


def beautify_master(path, sheet_name="분류"):
    """마스터 파일 서식 적용."""
    wb = load_workbook(path)
    ws = wb[sheet_name]
    center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center

    header_to_colidx = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

    def fmt(header, fmt_str):
        if header in header_to_colidx:
            c = header_to_colidx[header]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c).number_format = fmt_str

    fmt("정제 총 매출", "#,##0")
    fmt("평균 구매금", "#,##0.##")
    fmt("구매 횟수", "#,##0")
    fmt("가중치(점수)", "0")

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            text = "" if val is None else str(val)
            max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = max(10, min(60, int(max_len * 1.3)))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    if "가중치(점수)" in header_to_colidx:
        ws.column_dimensions[get_column_letter(header_to_colidx["가중치(점수)"])].hidden = True

    wb.save(path)


# ====== 메인 함수: run_modules.py 에서 호출할 함수 ======
def run_crm_scoring(base_name: str, input_dir: str = ".") -> None:
    """
    기존 단일 스크립트(crm_scoring.py)를 함수 형태로 감싼 버전.
    base_name(예: '20251117_업데이트')만 넘겨주면, 같은 폴더에서
    csv/xlsx를 찾아 4개 엑셀 파일(마스터 + 3명 타깃)을 생성한다.
    """
    input_dir_path = Path(input_dir)
    candidate_paths = [
        input_dir_path / f"{base_name}.csv",
        input_dir_path / f"{base_name}.xlsx",
    ]

    # ====== 입력 로드 ======
    df = None
    last_err = None
    for path in candidate_paths:
        if not os.path.exists(path):
            continue
        try:
            if path.suffix.lower() == ".csv":
                for enc in ["utf-8-sig", "cp949", "utf-8", "euc-kr"]:
                    try:
                        df = pd.read_csv(path, encoding=enc)
                        break
                    except Exception as e:
                        last_err = e
                if df is not None:
                    print(f"[crm_scoring] CSV 읽기 성공: {path}")
                    break
            else:
                df = pd.read_excel(path)
                print(f"[crm_scoring] Excel 읽기 성공: {path}")
                break
        except Exception as e:
            last_err = e

    if df is None:
        raise RuntimeError(
            f"[crm_scoring] 입력 파일을 찾거나 읽을 수 없습니다: {candidate_paths} / 마지막 오류: {last_err}"
        )

    df.columns = [str(c).strip() for c in df.columns]

    # ====== 컬럼 매핑 ======
    col_name = find_col(df.columns, [["환자", "명"], ["환자", "이름"], ["고객", "명"], ["성명"], ["이름"]])
    col_phone = find_col(df.columns, [["연락처"], ["전화"], ["휴대폰"], ["핸드폰"], ["모바일"]])
    col_total = find_col(df.columns, [["정제", "총", "매출"], ["정재", "총", "매출"], ["정제", "매출"], ["총", "매출"], ["매출", "합계"]])
    col_cnt = find_col(df.columns, [["구매", "횟수"], ["결제", "횟수"], ["방문", "횟수"], ["내원", "횟수"], ["횟수"]])
    col_avg = find_col(df.columns, [["평균", "구매", "금"], ["평균", "구매", "액"], ["평균", "결제", "액"], ["객단가"], ["평균", "단가"]])
    col_recent = find_col(
        df.columns,
        [
            ["최근", "진료", "일"],
            ["최근", "구매", "일"],
            ["최근", "방문", "일"],
            ["최근", "내원", "일"],
            ["최근", "결제", "일"],
            ["마지막", "방문", "일"],
        ],
    )
    col_first = find_col(
        df.columns,
        [
            ["최초", "구매", "일"],
            ["첫", "구매", "일"],
            ["첫", "방문", "일"],
            ["첫", "내원", "일"],
            ["첫", "결제", "일"],
        ],
    )

    required_cols = {
        "환자명": col_name,
        "연락처": col_phone,
        "정제 총 매출": col_total,
        "구매 횟수": col_cnt,
        "평균 구매금": col_avg,
        "최초 구매일": col_first,
        "최근 구매일": col_recent,
    }

    work = pd.DataFrame()
    for out_col, src_col in required_cols.items():
        work[out_col] = df[src_col] if (src_col is not None and src_col in df.columns) else np.nan

    # ====== 변환 ======
    work["최초 구매일"] = work["최초 구매일"].apply(to_date)
    work["최근 구매일"] = work["최근 구매일"].apply(to_date)
    for num_col in ["정제 총 매출", "구매 횟수", "평균 구매금"]:
        work[num_col] = work[num_col].apply(to_num)

    # ====== 2025년 최근일자 필터 ======
    work = work[work["최근 구매일"].apply(lambda d: (not pd.isna(d)) and d.year == 2025)].copy()

    # ====== 그룹/이름 보정 ======
    groups, fixed_names = [], []
    for _, row in work.iterrows():
        g, fixed = extract_group_anywhere(row["환자명"], row["최초 구매일"])
        groups.append(g)
        fixed_names.append(fixed)
    work["그룹"] = groups
    work["환자명"] = fixed_names

    # ====== 규칙 플래그 ======
    work["_over90"] = work["최근 구매일"].apply(lambda d: over_90_days(d, today))
    work["_total_is_zero"] = work["정제 총 매출"].apply(lambda x: (not pd.isna(x)) and float(x) == 0.0)

    # ====== X3 사전 확정 (강제3) ======
    work["_forced3"] = work["_over90"] | work["_total_is_zero"]

    # ====== 상관관계 기반 가중치 자동 산출 (비-강제3 전체 대상) ======
    mask_weight = (~work["_forced3"]) & work[["정제 총 매출", "구매 횟수", "평균 구매금"]].notna().all(axis=1)
    feat_df = work.loc[mask_weight, ["정제 총 매출", "구매 횟수", "평균 구매금"]].copy()

    auto_weights = compute_corr_weights(feat_df)
    sum_w = sum(auto_weights.values())

    # ====== 점수 계산: 강제3 제외 + 군별 robust(10~90%) 스케일 ======
    work["_score100"] = np.nan  # 강제3은 NaN 유지

    for g in ["A", "C", "D", "E"]:
        mask_group = work["그룹"] == g
        mask_calc = mask_group & (~work["_forced3"])
        if not mask_calc.any():
            continue

        sub = work.loc[mask_calc, ["정제 총 매출", "구매 횟수", "평균 구매금"]].astype(float)

        # 1) 군 내부 min-max 정규화
        norm = pd.DataFrame(index=sub.index)
        for col in sub.columns:
            cmin, cmax = sub[col].min(skipna=True), sub[col].max(skipna=True)
            if pd.isna(cmin) or pd.isna(cmax) or cmax == cmin:
                norm[col] = 0.5
            else:
                norm[col] = (sub[col] - cmin) / (cmax - cmin)

        # 2) 가중합
        score01 = (
            auto_weights["정제 총 매출"] * norm["정제 총 매출"]
            + auto_weights["구매 횟수"] * norm["구매 횟수"]
            + auto_weights["평균 구매금"] * norm["평균 구매금"]
        ) / sum_w

        # 3) robust 10~90% 재스케일
        q10 = float(np.quantile(score01, 0.10)) if len(score01) > 0 else 0.0
        q90 = float(np.quantile(score01, 0.90)) if len(score01) > 0 else 1.0
        if q90 <= q10:
            robust01 = pd.Series(0.5, index=score01.index)
        else:
            robust01 = ((score01 - q10) / (q90 - q10)).clip(0, 1)

        work.loc[mask_calc, "_score100"] = (robust01 * 100).round(0)

    # 표시/정렬용
    work["가중치"] = [weight_text(g, s) for g, s in zip(work["그룹"], work["_score100"])]
    work["가중치(점수)"] = work["_score100"]  # 정렬·필터용(숫자)

    # ====== 등급 산정: 65점 절대 기준만 사용 ======
    grades_abs = pd.Series(index=work.index, dtype="object")
    idx_calc_all = work.index[(~work["_forced3"]) & (pd.notna(work["_score100"]))]
    for i in idx_calc_all:
        g = work.at[i, "그룹"]
        grades_abs.at[i] = f"{g}1" if work.at[i, "_score100"] >= ABSOLUTE_SCORE_CUT else f"{g}2"

    # 강제3 일괄 부여
    for i in work.index[work["_forced3"]]:
        g = work.at[i, "그룹"]
        grades_abs.at[i] = f"{g}3"

    work["등급"] = grades_abs  # 최종 등급은 절대 65점 컷만 반영

    # ====== 공통 출력 컬럼 ======
    out_cols = [
        "환자명",
        "연락처",
        "정제 총 매출",
        "구매 횟수",
        "평균 구매금",
        "최초 구매일",
        "최근 구매일",
        "가중치",
        "가중치(점수)",
        "등급",
    ]

    # ====== 마스터 파일(절대 65점 컷 기준) ======
    master = work[out_cols].copy()
    master_name = str(Path(input_dir) / f"{file_prefix}_환자분류_절대65점컷기준.xlsx")

    with pd.ExcelWriter(master_name, engine="openpyxl") as w:
        master.to_excel(w, index=False, sheet_name="분류")

    # ====== 45~90일 범위 계산(최근 구매일 기준) ======
    work["_days_since"] = work["최근 구매일"].apply(lambda d: (today - d).days if pd.notna(d) else np.nan)
    mask_45_90 = work["_days_since"].between(45, 90, inclusive="both")

    # ====== 1) 김훈하 약사님 파일: 45~90일 & A군 & A1/A2 & 점수 50~100 ======
    mask_kim = (
        mask_45_90
        & (work["그룹"] == "A")
        & (work["등급"].isin(["A1", "A2"]))
        & (pd.notna(work["_score100"]))
        & (work["_score100"].between(50, 100, inclusive="both"))
    )
    kim_df = work.loc[mask_kim, out_cols].copy()
    kim_df.sort_values(by="가중치(점수)", ascending=False, inplace=True)
    kim_name = str(Path(input_dir) / f"{file_prefix}_김훈하_약사님.xlsx")

    with pd.ExcelWriter(kim_name, engine="openpyxl") as w:
        kim_df.to_excel(w, index=False, sheet_name="대상")

    # ====== 2) 전정미 약사님 파일: 45~90일 & C군 (C1/C2/C3 모두) ======
    mask_jeon = mask_45_90 & (work["그룹"] == "C")
    jeon_df = work.loc[mask_jeon, out_cols].copy()
    # 점수 NaN(C3 등)도 존재 가능 → NaN은 정렬에서 뒤로
    jeon_df["__sort"] = jeon_df["가중치(점수)"].fillna(-1e9)
    jeon_df.sort_values(by="__sort", ascending=False, inplace=True)
    jeon_df.drop(columns=["__sort"], inplace=True)
    jeon_name = str(Path(input_dir) / f"{file_prefix}_전정미_약사님.xlsx")

    with pd.ExcelWriter(jeon_name, engine="openpyxl") as w:
        jeon_df.to_excel(w, index=False, sheet_name="대상")

    # ====== 3) 백인보 약사님 파일: 45~90일 & (A군 30~<50점) OR (D/E군 ≥40점) ======
    mask_baek = mask_45_90 & (
        (
            (work["그룹"] == "A")
            & (pd.notna(work["_score100"]))
            & (work["_score100"] >= 30)
            & (work["_score100"] < 50)
            & (~work["_forced3"])
        )
        | (
            (work["그룹"].isin(["D", "E"]))
            & (pd.notna(work["_score100"]))
            & (work["_score100"] >= 40)
            & (~work["_forced3"])
        )
    )
    baek_df = work.loc[mask_baek, out_cols].copy()
    baek_df.sort_values(by="가중치(점수)", ascending=False, inplace=True)
    baek_name = str(Path(input_dir) / f"{file_prefix}_백인보_약사님.xlsx")

    with pd.ExcelWriter(baek_name, engine="openpyxl") as w:
        baek_df.to_excel(w, index=False, sheet_name="대상")

    # ====== 서식 적용 ======
    beautify_excel(kim_name, "대상")
    beautify_excel(jeon_name, "대상")
    beautify_excel(baek_name, "대상")
    beautify_master(master_name, "분류")

    # ====== 콘솔 요약 ======
    print("자동 산출된 가중치(합=10):")
    for k in ["정제 총 매출", "구매 횟수", "평균 구매금"]:
        print(f" - {k}: {auto_weights.get(k):.2f}")

    print("\n생성 파일:")
    print(f"- 마스터(절대65점 컷): {master_name}")
    print(f"- 김훈하: {kim_name}")
    print(f"- 전정미: {jeon_name}")
    print(f"- 백인보: {baek_name}")


# ====== 이 파일을 단독으로 실행했을 때: 최신 *_업데이트 자동 사용 ======
if __name__ == "__main__":
    base_dir = Path(".")
    candidates = sorted(
        [p for p in base_dir.glob("*_업데이트.*") if p.suffix.lower() in {".csv", ".xlsx"}]
    )
    if not candidates:
        print("[crm_scoring] *_업데이트.csv/.xlsx 파일을 찾지 못했습니다.")
    else:
        latest = candidates[-1]
        base_name = latest.stem  # 확장자 제거
        print(f"[crm_scoring] 최신 업데이트 파일 기준으로 실행: {latest}")
        run_crm_scoring(base_name=base_name, input_dir=".")
